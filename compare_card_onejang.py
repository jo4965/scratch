from openpyxl import load_workbook, Workbook
# import argparse
import xlrd
import glob
import os
import re
import bs4
# parser = argparse.ArgumentParser(description='Compare onejang vs card')
# parser.add_argument('--card-files', nargs='+',
#                     help='카드내역서 엑셀 파일 (복수 가능)')

# parser.add_argument('--onejang-file', help='원장 엑셀 파일 (복수 가능)')

# args = parser.parse_args()


if __name__ == '__main__':
    card_files = glob.glob(os.path.dirname(os.path.abspath(__file__))+ "/card*")
    print(os.path.dirname(os.path.abspath(__file__)) + "/onejang*")
    onejang_file = glob.glob(os.path.dirname(os.path.abspath(__file__)) + "/onejang*")
    print(onejang_file)
    onejang_file = onejang_file[0]

    # onejang_file = "/Users/user/Downloads/onejang_old.xlsx"
    card_money_to_rows = dict()
    
    for card_file_path in card_files:
        try:
            is_card_xlrd = 0
            # data_only=True로 해줘야 수식이 아닌 값으로 받아온다. 
            card_wb = load_workbook(card_file_path, data_only=True)
            # 시트 이름으로 불러오기 
            card_ws = card_wb['법인매출상세내역']

            # 모든 행 단위로 출력

            for row in card_ws.rows:
                if row[0].value == '카드번호':
                    card_head_row = row
                    continue
                
                # for cell in row:
                #     print(cell.value, end=' ')
                # print('')

                money = row[4].value
                if money in card_money_to_rows.keys():
                    card_money_to_rows[money].append(row)
                else:
                    card_money_to_rows[money] = [row]
        except:
            try:
                print("load_workbook error: {}".format(card_file_path))
                is_card_xlrd = 1
                card_wb = xlrd.open_workbook(card_file_path)
                card_ws = card_wb.sheet_by_index(0)

                for i in range(card_ws.nrows):
                    if card_ws.row_values(i)[0] == "카드번호":
                        card_head_row = row
                        continue

                    row = card_ws.row_values(i)
                    money = card_ws.row_values(i)[4]
                    # if not isinstance(money, float):
                    #     print("not isinstance money {}".format(money))
                    #     continue

                    # money = int(row[22])

                    if money in card_money_to_rows.keys():
                        card_money_to_rows[money].append(row)
                    else:
                        card_money_to_rows[money] = [row]
            except:
                is_card_xlrd = 2
                print("xlrd.open_workbook error {}".format(card_file_path))
                with open(card_file_path, "rt") as fp:
                    soup = bs4.BeautifulSoup(fp.read(), 'html.parser')
                    # print(soup)
                    head_row = soup.findAll("th", {"style" : "background-color:#D9D9D9;"})
                    card_head_row = list()
                    for th in head_row:
                        card_head_row.append(th.text)

                    row_finder = soup.findAll("td")
                    # money_row = list()
                    money_row = list()
                    for i, td in enumerate(row_finder):
                        # money_row.append(th.text)
                        money_row.append(td.text)
                        if (i+1) % 14 == 0:
                            money = int(money_row[4])
                            if money in card_money_to_rows.keys():
                                card_money_to_rows[money].append(money_row)
                            else:
                                card_money_to_rows[money] = [money_row]
                            money_row = list()




    onejang_money_to_rows = dict()

    try:
        onejang_wb = xlrd.open_workbook(onejang_file)
        onejang_ws = onejang_wb.sheet_by_index(0)

        for i in range(onejang_ws.nrows):
            if onejang_ws.row_values(i)[1].strip() == "":
                continue

            if onejang_ws.row_values(i)[22] == "":
                continue

            if onejang_ws.row_values(i)[22] == "대변금액":
                onejang_head_row = onejang_ws.row_values(i)
                continue
            
            row = onejang_ws.row_values(i)

            if not isinstance(row[22], float):
                continue

            money = int(row[22])
            if money == 0:
                continue


            if money in onejang_money_to_rows.keys():
                onejang_money_to_rows[money].append(row)
            else:
                onejang_money_to_rows[money] = [row]

            # for j in range(onejang_ws.ncols):
            #     print(onejang_ws.row_values(i)[j], end=' ')
        is_onejang_xlrd = True
            # print("")
    except Exception as e:
        # data_only=True로 해줘야 수식이 아닌 값으로 받아온다. 
        onejang_wb = load_workbook(onejang_file, data_only=True)
        # 시트 이름으로 불러오기 
        onejang_ws = onejang_wb['Sheet1']

        for row in onejang_ws.rows:

            # for cell in row:
            #     print(cell.value, end=' ')
            # print('')

            if row[1].value == None:
                continue
            
            if row[1].value.strip() == "":
                continue

            if row[22].value == "":
                continue


            if row[22].value == '대변금액':
                onejang_head_row = row
                continue

            money = row[22].value

            if money == None:
                continue

            if not isinstance(money, int):
                continue

            if money in onejang_money_to_rows.keys():
                onejang_money_to_rows[money].append(row)
            else:
                onejang_money_to_rows[money] = [row]
        is_onejang_xlrd = False
        # onejang_wb = xlrd.open_workbook(onejang_file)
        # onejang_ws = onejang_wb.sheet_by_index(0)
    
    card_but_not_in_onejang = list()
    print("[+] 카드 내역에는 있는데 원장에 없는 놈들 확인\n")
    if not is_card_xlrd:
        for card_money, card_rows in card_money_to_rows.items():
            
            if card_money not in onejang_money_to_rows.keys():
                
                for row in card_rows:
                    print("[+] 원장에 없어요 -> ", end = ' ')
                    for cell in row:
                        print(cell.value, end=' ')
                    print('')

                    card_but_not_in_onejang.append(row)
                print('')
            else:
                if len(card_rows) > len(onejang_money_to_rows[card_money]):
                    not_exist_cnt = len(card_rows) - len(onejang_money_to_rows[card_money])
                    for i in range(not_exist_cnt):
                        print("[+] 원장에 없어요 -> ", end = ' ')
                        for cell in card_rows[i]:
                            print(cell.value, end=' ')
                        print('')
                        card_but_not_in_onejang.append(card_rows[i])

                    print('')
                    
                    # print("[+] 얘네 중에 원장에 없는 애들이 있어요")
                    # for row in card_rows:
                    #     print("---->", end = ' ')
                    #     for cell in row:
                    #         print(cell.value, end=' ')
                    #     print('')
                    # print('')    

    else:
        for card_money, card_rows in card_money_to_rows.items():
            
            if card_money not in onejang_money_to_rows.keys():
                
                for row in card_rows:
                    print("[+] 원장에 없어요 -> ", end = ' ')
                    for value in row:
                        print(value, end=' ')
                    print('')

                    card_but_not_in_onejang.append(row)
                print('')
            else:
                if len(card_rows) > len(onejang_money_to_rows[card_money]):
                    not_exist_cnt = len(card_rows) - len(onejang_money_to_rows[card_money])
                    for i in range(not_exist_cnt):
                        print("[+] 원장에 없어요 -> ", end = ' ')
                        for value in card_rows[i]:
                            print(value, end=' ')
                        print('')
                        card_but_not_in_onejang.append(card_rows[i])

                    print('')


    print("\n\n")
    print("======================================")
    print("======================================")

    onejang_but_not_in_card = list()
    
    print("[+] 원장에는 있는데 카드내역에 없는 놈들 확인\n")

    if not is_onejang_xlrd:
        for onejang_money, onejang_rows in onejang_money_to_rows.items():
            
            if onejang_money not in card_money_to_rows.keys():            
                for row in onejang_rows:
                    print("[+] 카드 내역에 없어요 -> ", end = ' ')
                    for cell in row:
                        print(cell.value, end=' ')
                    print('')

                    onejang_but_not_in_card.append(row)
                print('')
            else:
                if len(onejang_rows) > len(card_money_to_rows[onejang_money]):
                    not_exist_cnt = len(onejang_rows) - len(card_money_to_rows[onejang_money])
                    for i in range(not_exist_cnt):
                        print("[+] 카드 내역에 없어요 -> ", end = ' ')
                        for cell in onejang_rows[i]:
                            print(cell.value, end=' ')
                        print('')
                        onejang_but_not_in_card.append(onejang_rows[i])
                    print('')
                    # print("[+] 얘네 중에 카드 내역에 없는 애들이 있어요")
                    # for row in onejang_rows:
                    #     print("---->", end = ' ')
                    #     for value in row:
                    #         print(value, end=' ')
                    #     print('')
                    # print('')
    else:
        for onejang_money, onejang_rows in onejang_money_to_rows.items():
                    
            if onejang_money not in card_money_to_rows.keys():            
                for row in onejang_rows:
                    print("[+] 카드 내역에 없어요 -> ", end = ' ')
                    for value in row:
                        print(value, end=' ')
                    print('')
                    onejang_but_not_in_card.append(row)
                print('')
            else:
                if len(onejang_rows) > len(card_money_to_rows[onejang_money]):
                    not_exist_cnt = len(onejang_rows) - len(card_money_to_rows[onejang_money])
                    for i in range(not_exist_cnt):
                        print("[+] 카드 내역에 없어요 -> ", end = ' ')
                        for value in onejang_rows[i]:
                            print(value, end=' ')
                        print('')
                        onejang_but_not_in_card.append(onejang_rows[i])
                    print('')



    new_workbook = Workbook()
    card_sheet = new_workbook.active
    card_sheet.title = "카드 내역에는 있는데 원장에 없는 놈들"

    if not is_card_xlrd:
        for head_col_num in range(len(card_head_row)):
            card_sheet.cell(row = 1, column = head_col_num + 1).value = card_head_row[head_col_num].value

        for row_idx in range(0, len(card_but_not_in_onejang)):
            for col_idx in range(len(card_but_not_in_onejang[row_idx])):
                card_sheet.cell(row = row_idx+2, column = col_idx + 1).value = card_but_not_in_onejang[row_idx][col_idx].value
    else:
        for head_col_num in range(len(card_head_row)):
            card_sheet.cell(row = 1, column = head_col_num + 1).value = card_head_row[head_col_num]

        for row_idx in range(0, len(card_but_not_in_onejang)):
            for col_idx in range(len(card_but_not_in_onejang[row_idx])):
                card_sheet.cell(row = row_idx+2, column = col_idx + 1).value = card_but_not_in_onejang[row_idx][col_idx]



    onejang_sheet = new_workbook.create_sheet()
    onejang_sheet.title = "원장에는 있는데 카드 내역에 없는 놈들"

    if not is_onejang_xlrd:
        for head_col_num in range(len(onejang_head_row)):
            onejang_sheet.cell(row = 1, column = head_col_num + 1).value = onejang_head_row[head_col_num].value


        for row_idx in range(0, len(onejang_but_not_in_card)):
            for col_idx in range(len(onejang_but_not_in_card[row_idx])):
                onejang_sheet.cell(row = row_idx+2, column = col_idx + 1).value = onejang_but_not_in_card[row_idx][col_idx].value
    else:
        for head_col_num in range(len(onejang_head_row)):
            onejang_sheet.cell(row = 1, column = head_col_num + 1).value = onejang_head_row[head_col_num]

        for row_idx in range(0, len(onejang_but_not_in_card)):
            for col_idx in range(len(onejang_but_not_in_card[row_idx])):
                onejang_sheet.cell(row = row_idx+2, column = col_idx + 1).value = onejang_but_not_in_card[row_idx][col_idx]

    new_workbook.save("./result.xlsx")




